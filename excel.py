from openpyxl import load_workbook, Workbook
import plotly.graph_objects as go
import re


class Excel:

    def __init__(self):
        self.workbook = None
        self.workbook_name = None
        self.worksheet = None

    def open_workbook(self, workbook_name: str):
        '''
        Open a .xlsx format workbook for reading and writing.
        :param workbook_name: Name of a workbook, does not have to include the extension
        :return: None
        '''

        # check if the extension was included in the workbook name.
        if ".xlsx" not in workbook_name:
            # if not, append it
            workbook_name += ".xlsx"

        # remember what the workbook is called for when we save and close.
        self.workbook_name = workbook_name

        try:
            # try to open the workbook
            self.workbook = load_workbook(filename=workbook_name)
        except FileNotFoundError:
            # if the file was not found, create the workbook. This does not open the workbook
            self.create_workbook(workbook_name)
            try:
                # open the newly created workbook
                self.workbook = load_workbook(filename=workbook_name)
            except FileNotFoundError:
                # If the file was still not found, it could not be created.
                raise (NoWorkbookFound(400))

        # set the worksheet to the active sheet in the workbook
        self.worksheet = self.workbook.active

        return None

    def save_and_close(self, new_name=None):
        """
        Save and close the open workbook
        :param new_name: Optional new name of workbook
        :return: None
        """
        # if a new name for the workbook was not given
        if new_name is None:
            # save the workbook with the saved name
            self.workbook.save(self.workbook_name)

        else:
            # check if the extension was included in the workbook name.
            if ".xlsx" not in new_name:
                # if not, append it
                new_name += ".xlsx"
            # save the workbook with the new name
            self.workbook.save(new_name)

        return None

    def read_cell(self, cell: str):
        """
        Read a single cell's value
        :param cell: String address of the cell to read
        :return: Integer value of cell if cell contains an Integer, else a String of the cell's value
        """

        # check that the workbook is open
        if self.workbook is None or self.worksheet is None:
            raise NoWorkbookFound(401)

        # check that the given address is valid
        if self.check_address(cell) is False:
            raise InvalidCellAddress(201)

        # get the value from the sheet
        sheet_value = self.worksheet[cell].value

        # check that a value was retrieved
        if sheet_value == "":
            raise NoValueFound(101)
        if sheet_value is None:
            raise NoValueFound(102)

        try:
            # try to return an integer of the value
            return int(sheet_value)
        except ValueError:
            # else return it as a string
            return sheet_value

    def write_cell(self, cell: str, value):
        """
        Write a new value to a single cell
        :param cell: String address of the cell to be written
        :param value: Typeless new value to be written to cell.
        :return: None
        """
        # check that the workbook is open
        if self.workbook is None or self.worksheet is None:
            raise NoWorkbookFound(402)

        # check that a valid address was given
        if self.check_address(cell) is False:
            raise InvalidCellAddress(202)

        # write the value to the cell
        self.worksheet[cell] = value

        return None

    '''
    def create_graph(self, col_a: str, col_b: str, title: str, y_only_mode=False, a_name=None, b_name=None):
        """
        Create and display a graph of the given data. This function has two modes: "x,y" and "y only". x,y mode takes
        the first range as all the x values and the second range as all the y values. y only mode takes each range
        as separate data points, with the x values for each line being x for index in length of range.
        :param col_a: String Excel style cell range i.e. "A1:A4". Must be a single column.
        :param col_b: String Excel style cell range i.e. "A1:A4". Must be a single column.
        :param title: String A title for the graph
        :param y_only_mode: Optional Boolean denoting which mode the function should use.
        :param a_name: Optional Name for the first line. Only used in "y only" mode.
        :param b_name: Optional Name for the second line. Only used in "y only" mode.
        :return: None
        """
        # check that the workbook is open
        if self.workbook is None or self.worksheet is None:
            raise NoWorkbookFound(403)

        # declare variables to use in this function
        x_vals = []
        y_vals = []
        x_ranges = []
        y_ranges = []
        fig = go.Figure()

        # Next block of code checks for valid input ranges
        # simple check that two cell addresses were given
        if ":" not in col_a or ":" not in col_b:
            raise InvalidCellAddress(204)
        # split those addresses apart
        x_ranges = col_a.split(":")
        y_ranges = col_b.split(":")
        # check the individual addresses are valid
        for cell_addr in x_ranges + y_ranges:
            if self.check_address(cell_addr) is False:
                raise InvalidCellAddress(203)

        # check that each range is a single column and that the second cell is lower than the first.
        if re.search("([A-Z]+)\d+", x_ranges[0]).group(1) != re.search("([A-Z]+)\d+", x_ranges[1]).group(1) or \
                re.search("([A-Z]+)\d+", y_ranges[0]).group(1) != re.search("([A-Z]+)\d+", y_ranges[1]).group(1) or \
                re.search("[A-Z]+(\d+)", x_ranges[0]).group(1) > re.search("[A-Z]+(\d+)", x_ranges[1]).group(1) or \
                re.search("[A-Z]+(\d+)", y_ranges[0]).group(1) > re.search("[A-Z]+(\d+)", y_ranges[1]).group(1):
            raise InvalidCellAddress(205)

        if y_only_mode:

            # empty list for the first line
            y = []
            # get the column letter
            cell_letter = re.search("([A-Z]+)\d+", x_ranges[0]).group(1)
            # get the first row number
            cell_num = int(re.search("[A-Z]+(\d+)", x_ranges[0]).group(1))
            # get the last row number
            max_num = int(re.search("[A-Z]+(\d+)", x_ranges[1]).group(1))
            # iter value will be used to set the x values for this line
            iter = 1
            # while still working down the column
            while cell_num <= max_num:
                # append the next value in the sheet to known y values
                y.append(self.read_cell(cell_letter + str(cell_num)))
                # append the next iter val to the known x values
                x_vals.append(iter)
                # increment the row
                cell_num += 1
                iter += 1
            # add the created line to the figure
            fig.add_trace(go.Scatter(x=x_vals, y=y, name=a_name))

            # empty list for the second line
            y = []
            # get the column letter
            cell_letter = re.search("([A-Z]+)\d+", y_ranges[0]).group(1)
            # get the first row number
            cell_num = int(re.search("[A-Z]+(\d+)", y_ranges[0]).group(1))
            # get the last row number
            max_num = int(re.search("[A-Z]+(\d+)", y_ranges[1]).group(1))
            # iter value will be used to set the x values for this line
            iter = 1
            # while still working down the column
            while cell_num <= max_num:
                # append the next value in the sheet to known y values
                y.append(self.read_cell(cell_letter + str(cell_num)))
                # append the next iter val to the known x values
                x_vals.append(iter)
                # increment the row
                iter += 1
                cell_num += 1
            # add the created line to the figure
            fig.add_trace(go.Scatter(x=x_vals, y=y, name=b_name))

        # else using x,y mode
        else:

            # get the letter of the first column
            cell_letter = re.search("([A-Z]+)\d+", x_ranges[0]).group(1)
            # get the first row number
            cell_num = int(re.search("[A-Z]+(\d+)", x_ranges[0]).group(1))
            # get the last row number
            max_num = int(re.search("[A-Z]+(\d+)", x_ranges[1]).group(1))
            # while still working down the column
            while cell_num <= max_num:
                # append each val in the column to the x_vals
                x_vals.append(self.read_cell(cell_letter + str(cell_num)))
                # increment the row
                cell_num += 1

            # get the letter of the second column
            cell_letter = re.search("([A-Z]+)\d+", y_ranges[0]).group(1)
            # get the first row number
            cell_num = int(re.search("[A-Z]+(\d+)", y_ranges[0]).group(1))
            # get the last row number
            max_num = int(re.search("[A-Z]+(\d+)", y_ranges[1]).group(1))
            # while still working down the column
            while cell_num <= max_num:
                # append each value in the column to known y values
                y_vals.append(self.read_cell(cell_letter + str(cell_num)))
                # increment row
                cell_num += 1

            # add the created line to the figure
            fig.add_trace(go.Scatter(x=x_vals, y=y_vals))

        # add the given title to the figure
        fig.update_layout(title=go.layout.Title(
            text=title,
            xref="paper",
            x=0))
        # show the figure
        fig.show()

        return None
    '''

    @staticmethod
    def create_workbook(workbook_name):
        """
        Create a new workbook
        :param workbook_name: Name of new workbook
        :return: None
        """
        # create a new workbook object
        workbook = Workbook()
        # save it with the given name
        workbook.save(workbook_name)
        return None

    @staticmethod
    def check_address(cell_address: str):
        """
        Checks the validity of a cell address. Cell addresses must be in the form [A-Z]+\d+.
        :param cell_address: String address of a cell
        :return: Boolean. True is cell address is valid else false
        """
        try:
            # check that the valid format is being used
            re.search("([A-Z]+\d+)", cell_address).group(1)
        except AttributeError:
            # if not, return false
            return False
        return True


class NoWorkbookFound(Exception):
    # 400 No Workbook found and a workbook could not be created.
    # 401 Error occurred while reading.
    # 402 Error occurred while writing.
    # 403 Error occurred while creating a graph.
    def __init__(self, code):
        self.code = code

    def __str__(self):
        return repr(self.code)


class NoValueFound(Exception):
    # 101 Value is an empty string.
    # 102 Value is NoneType.
    def __init__(self, code):
        self.code = code

    def __str__(self):
        return repr(self.code)


class InvalidCellAddress(Exception):
    # 201 Invalid Address given while reading.
    # 202 Invalid Address given while writing.
    # 203 Invalid Address given while graphing.
    # 204 Invalid Range given while graphing.
    # 205 X ranges or Y ranges do not match.
    def __init__(self, code):
        self.code = code

    def __str__(self):
        return repr(self.code)
